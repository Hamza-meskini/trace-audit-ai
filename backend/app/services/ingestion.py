"""Document ingestion service — parse PDF, DOCX, XLSX files into text chunks."""

import re
from pathlib import Path
from dataclasses import dataclass


@dataclass
class ParsedChunk:
    """A chunk of text extracted from a document with metadata."""
    content: str
    page_number: int | None
    chunk_index: int
    metadata: dict | None = None


def parse_pdf(file_path: str) -> list[ParsedChunk]:
    """Extract text from a PDF file, one chunk per page."""
    import fitz  # PyMuPDF

    chunks = []
    doc = fitz.open(file_path)
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text().strip()
        if text:
            chunks.append(ParsedChunk(
                content=text,
                page_number=page_num + 1,
                chunk_index=page_num,
            ))
    doc.close()
    return chunks


def parse_docx(file_path: str) -> list[ParsedChunk]:
    """Extract text from a DOCX file, grouped by heading sections or requirement boundaries."""
    from docx import Document

    doc = Document(file_path)
    chunks = []
    current_section = []
    chunk_idx = 0

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        is_heading = para.style and para.style.name.startswith("Heading")
        is_req_header = bool(re.match(r"^(REQ[-_]?[A-Za-z0-9_-]*\d+|R[-_]?[A-Za-z0-9_-]*\d+)\s*[:\-–]", text, re.IGNORECASE))

        # Start a new chunk on headings or requirement boundaries or length overflow
        if (is_heading or is_req_header or len("\n".join(current_section)) > 1500) and current_section:
            chunks.append(ParsedChunk(
                content="\n".join(current_section),
                page_number=None,
                chunk_index=chunk_idx,
            ))
            chunk_idx += 1
            current_section = []

        current_section.append(text)

    # Flush remaining
    if current_section:
        chunks.append(ParsedChunk(
            content="\n".join(current_section),
            page_number=None,
            chunk_index=chunk_idx,
        ))

    return chunks


def parse_xlsx(file_path: str) -> list[ParsedChunk]:
    """Extract text from an XLSX file, one chunk per row with headers as context."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    chunks = []
    chunk_idx = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Use first row as headers
        headers = [str(h) if h else f"Col{i}" for i, h in enumerate(rows[0])]

        for row_idx, row in enumerate(rows[1:], start=2):
            cells = [str(c) if c is not None else "" for c in row]
            if not any(cells):
                continue

            # Build readable text: "Header1: Value1; Header2: Value2; ..."
            pairs = [f"{h}: {v}" for h, v in zip(headers, cells) if v]
            text = "; ".join(pairs)

            chunks.append(ParsedChunk(
                content=text,
                page_number=None,
                chunk_index=chunk_idx,
                metadata={"sheet": sheet, "row": row_idx},
            ))
            chunk_idx += 1

    wb.close()
    return chunks


def parse_csv(file_path: str) -> list[ParsedChunk]:
    """Extract text from a CSV file, one chunk per row."""
    import csv

    chunks = []
    chunk_idx = 0

    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, start=2):
            pairs = [f"{k}: {v}" for k, v in row.items() if v]
            if not pairs:
                continue
            text = "; ".join(pairs)
            chunks.append(ParsedChunk(
                content=text,
                page_number=None,
                chunk_index=chunk_idx,
                metadata={"row": row_idx},
            ))
            chunk_idx += 1

    return chunks


def parse_document(file_path: str) -> list[ParsedChunk]:
    """Route to the appropriate parser based on file extension."""
    ext = Path(file_path).suffix.lower()
    parsers = {
        ".pdf": parse_pdf,
        ".docx": parse_docx,
        ".xlsx": parse_xlsx,
        ".csv": parse_csv,
    }
    parser = parsers.get(ext)
    if not parser:
        raise ValueError(f"Unsupported file type: {ext}")
    return parser(file_path)
